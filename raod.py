import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

file_name = "Python_SDET_2_Week_Full_Ultra_Detailed.xlsx"

roadmap = []

# -------------------------------
# Week 1 - Saturday - Python Basics (Ultra-Detailed)
# -------------------------------
# Variables & Assignment
variables = ["Declare int variable","Declare float variable","Declare string variable","Declare bool variable","Declare complex variable",
             "Assign value to variable","Swap variables using temp","Swap variables without temp","Multiple assignment"]
# Strings - each method broken into multiple sub-tasks
string_methods = ["str.upper()","str.lower()","str.title()","str.capitalize()",
                  "str.split() on whitespace","str.split() on comma","str.join() with list","str.replace() single char",
                  "str.replace() substring","str.replace() regex","str.find()","str.count()","str.format() basic",
                  "str.format() with multiple placeholders","f-string simple","f-string with expressions"]
# Operators - Arithmetic, Logical, Relational, Assignment
operators = ["Arithmetic +","Arithmetic -","Arithmetic *","Arithmetic /","Arithmetic %","Arithmetic **","Arithmetic //",
             "Logical and","Logical or","Logical not",
             "Relational ==","Relational !=","Relational >","Relational <","Relational >=","Relational <=",
             "Assignment =","Assignment +=","Assignment -=","Assignment *=","Assignment /=","Ternary operator"]

# Loops
loops = ["for loop over list","for loop over tuple","for loop over dict keys","for loop with enumerate",
         "nested for loop","while loop","while loop with break","while loop with continue","do while simulation using while",
         "else clause with for loop","else clause with while loop"]

# Lists
lists = ["Create empty list","Create list with values","Nested list","Access element by index","Access element by negative index",
         "Slice list start:end","Slice list with step","Append element","Extend list","Insert element at index",
         "Remove element by value","Pop element by index","Clear list","Index() method","Count() method",
         "Sort ascending","Sort descending","Reverse list","Iterate using for loop","Iterate using while loop"]

# Tuples, Sets, Dicts
tuples = ["Create tuple","Access tuple element","Immutable property"]
sets = ["Create set","Add element","Remove element","Union","Intersection","Difference","Iterate set"]
dicts = ["Create dict","Access key","Access value","Add key-value","Update value","Remove key","Iterate keys","Iterate values","Iterate items","get() method","clear()"]

# Functions
functions = ["Define function with no args","Define function with args","Define function with kwargs",
             "Return single value","Return multiple values","Lambda function simple","Lambda function with args"]

# Modules
modules = ["import standard module","import specific function from module","pip install package","import installed package"]

# Combine Week 1 Saturday tasks
week1_sat_tasks = variables + string_methods + operators + loops + lists + tuples + sets + dicts + functions + modules

for task in week1_sat_tasks:
    roadmap.append(("Week 1","Saturday","Python Basics","Micro Task",task,0.25,"Not Started"))

# Week 1 Saturday - OOP (Ultra-detailed)
oop_tasks = ["Define class","Create object","Constructor __init__","Instance method","Class method @classmethod","Static method @staticmethod",
             "Private variable","Getter method","Setter method","Single inheritance","Multiple inheritance","Multilevel inheritance",
             "Hierarchical inheritance","Method overriding","Operator overloading +","Operator overloading -","Operator overloading *",
             "Operator overloading /","Abstract class","Abstract method","Encapsulation concept","Polymorphism: compile-time",
             "Polymorphism: runtime","Custom decorator","try block","except block","else block","finally block","Raise exception",
             "Custom exception"]

for task in oop_tasks:
    roadmap.append(("Week 1","Saturday","OOP","Micro Task",task,0.25,"Not Started"))

# Week 1 Sunday - Selenium + PyTest + Mini Project
selenium_tasks = ["Install Selenium","Setup WebDriver","Locator by id","Locator by name","Locator by class","Locator by absolute xpath",
                  "Locator by relative xpath","Locator by cssSelector","Locator by linkText","Locator by partialLinkText",
                  "click()","double click","right click","hover using ActionChains","send_keys()","clear()","submit()",
                  "Implicit wait","Explicit wait","Fluent wait","Handle alerts: accept","Handle alerts: dismiss","Handle alerts: send_keys",
                  "Switch frame by index","Switch frame by name","Switch frame by element","Switch window by handle","Switch window by title",
                  "Execute JavaScript","Scroll page","Take screenshot"]

pytest_tasks = ["Install pytest","Create test file","Run test via command line","Fixture: function scope","Fixture: class scope",
                "Fixture: module scope","Fixture: session scope","Autouse fixture","assertEqual","assertNotEqual","assertTrue","assertFalse",
                "assertIn","assertNotIn","assertRaises","parametrize","skip","xfail","Custom marker","HTML report","Allure report",
                "Hooks before_module","Hooks after_module","Hooks before_class","Hooks after_class"]

mini_project_tasks = ["Login automation script","Form filling script","Form submission script","Validation assertions","Generate report"]

week1_sun_tasks = selenium_tasks + pytest_tasks + mini_project_tasks

for task in week1_sun_tasks:
    roadmap.append(("Week 1","Sunday","Selenium/PyTest/Mini Project","Micro Task",task,0.5,"Not Started"))

# Week 2 Saturday - API + DB + Framework
api_tasks = ["GET simple","GET with params","GET with headers","GET with auth","POST simple","POST with payload","POST with headers","POST with auth",
             "PUT simple","PUT with payload","DELETE simple","DELETE with payload","Status code validation","Response time validation",
             "JSON key validation","JSON nested key validation","JSON schema validation","Auth Basic","Auth Token","Auth OAuth2",
             "Parameterized API: payload","Parameterized API: headers"]

db_tasks = ["Connect SQLite","Connect MySQL","Connect Postgres","INSERT single row","INSERT multiple rows","SELECT single row",
            "SELECT multiple rows","UPDATE single row","UPDATE multiple rows","DELETE single row","DELETE multiple rows",
            "Validate INSERT","Validate UPDATE","Validate DELETE","Commit transaction","Rollback transaction","Exception handling"]

framework_tasks = ["POM: create page class","POM: declare locators","POM: write methods","Data-driven: read CSV","Data-driven: read Excel",
                   "Data-driven: read JSON","Cross-browser Chrome","Cross-browser Firefox","Cross-browser Edge","Dynamic element handling",
                   "Logging setup","Log INFO","Log DEBUG","Log ERROR"]

week2_sat_tasks = api_tasks + db_tasks + framework_tasks

for task in week2_sat_tasks:
    roadmap.append(("Week 2","Saturday","API/DB/Framework","Micro Task",task,0.25,"Not Started"))

# Week 2 Sunday - Performance + Advanced PyTest + CI/CD + Capstone
performance_tasks = ["Install LocustIO","Write load test function","Parameterize load test","Run load test",
                     "Monitor response times","Monitor requests/sec","Monitor failures"]

advanced_pytest_tasks = ["Parallel execution pytest-xdist","Hooks/listeners","HTML report","Allure report"]

git_docker_tasks = ["Git init","Git commit","Git push","Git pull","Git branch","Git merge","Docker build image","Docker run container",
                    "Dockerfile FROM","Dockerfile COPY","Dockerfile RUN","Dockerfile CMD","Dockerfile ENV","Dockerfile WORKDIR"]

ci_cd_tasks = ["Jenkins create job","Jenkins add pipeline script","Jenkins trigger build","Integrate Git","Generate test reports"]

capstone_tasks = ["Project setup","Folder structure","Implement Selenium scripts","Implement API tests","Integrate DB validation",
                  "Generate reports","Parallel execution","Finalize project"]

week2_sun_tasks = performance_tasks + advanced_pytest_tasks + git_docker_tasks + ci_cd_tasks + capstone_tasks

for task in week2_sun_tasks:
    roadmap.append(("Week 2","Sunday","Performance/Advanced PyTest/Git/CI/CD/Capstone","Micro Task",task,0.5,"Not Started"))

# -------------------------------
# Create DataFrame & Excel
# -------------------------------
df = pd.DataFrame(roadmap, columns=["Week","Day","Topic","Subtopic","Task","Hours","Status"])
df["Progress %"] = ""
df.to_excel(file_name, index=False)

# Open workbook for dropdowns & progress formulas
wb = load_workbook(file_name)
ws = wb.active
ws.title = "Python_SDET_Full_Ultra_Detailed"

dv = DataValidation(type="list", formula1='"Not Started,In Progress,Done"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"G2:G{ws.max_row}")

for row in range(2, ws.max_row+1):
    ws[f"H{row}"] = f'=IF(G{row}="Done",100,IF(G{row}="In Progress",50,0))'

# Summary Sheet
if "Summary" in wb.sheetnames:
    wb.remove(wb["Summary"])
summary = wb.create_sheet("Summary")
summary["A1"] = "Python SDET 2-Week Ultra-Ultra-Detailed Progress Summary"
summary["A3"] = "Total Tasks"
summary["B3"] = f"=COUNTA(Python_SDET_Full_Ultra_Detailed!A2:A{ws.max_row})"
summary["A4"] = "Not Started"
summary["B4"] = f'=COUNTIF(Python_SDET_Full_Ultra_Detailed!G2:G{ws.max_row},"Not Started")'
summary["A5"] = "In Progress"
summary["B5"] = f'=COUNTIF(Python_SDET_Full_Ultra_Detailed!G2:G{ws.max_row},"In Progress")'
summary["A6"] = "Done"
summary["B6"] = f'=COUNTIF(Python_SDET_Full_Ultra_Detailed!G2:G{ws.max_row},"Done")'
summary["A7"] = "Overall Completion %"
summary["B7"] = f"=ROUND(AVERAGE(Python_SDET_Full_Ultra_Detailed!H2:H{ws.max_row}),2)"

wb.save(file_name)
print(f"✅ Fully Ultra-Ultra-Detailed Python SDET 2-Week Roadmap saved as {file_name}")
